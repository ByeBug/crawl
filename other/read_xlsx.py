# -*- coding: utf-8 -*-
"""
读取excel中的数据，并保存到数据库
Created on Thu Jun  7 16:48:57 2018

@author: storm
"""

import glob
import hashlib
import pymysql
from openpyxl import load_workbook


def read_bond_info(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    bond_infos = []
    bond_mains = []
    for i in ws.iter_rows(min_row=2):
        if i[0].value:
            s_id = i[0].value
            s_name = i[3].value
            s_short_name = i[1].value
            b_short_name = i[2].value
            init_face_value = i[4].value
            toal = i[5].value
            rest = i[6].value
            inter_begin_date = i[7].value
            inter_end_date = i[8].value
            honour_date = i[9].value
            deadline = i[10].value
            inter_type = i[11].value
            inter_rate = i[12].value
            inter_instr = i[13].value
            inter_bench = i[14].value
            type_bench = i[15].value
            inter_count = i[16].value
            list_date = i[17].value
            list_address = i[18].value
            trad_plat = i[19].value
            currency = i[20].value
            pay_order = i[21].value
            classify = i[22].value
            sfc = i[23].value
            lead_underwriter = i[24].value
            vice_lead_underwriter = i[25].value
            book_runner = i[26].value
            debt_subject = i[27].value
            buss_id = i[28].value
            spv = i[40].value
            original_owner = i[41].value
            basic_asset_type = i[42].value

            bond_info_item = (s_id, s_name, s_short_name, b_short_name, init_face_value, 
                toal, rest, inter_begin_date, inter_end_date, honour_date, deadline, inter_type, 
                inter_rate, inter_instr, inter_bench, type_bench, inter_count, list_date, list_address, 
                trad_plat, currency, pay_order, classify, sfc, lead_underwriter, vice_lead_underwriter, 
                book_runner, debt_subject, buss_id, spv, original_owner, basic_asset_type)

            bond_infos.append(bond_info_item)

            if i[27].value:
                c_id = hashlib.md5(i[27].value.encode('utf-8')).hexdigest()
                name = i[27].value
                is_listed = i[38].value
                uscc = i[28].value
                organ_num = i[29].value
                legal_name = i[30].value
                money = i[31].value
                busin_scope = i[34].value
                reg_address = i[37].value
                abstract = i[35].value
                main_bussiness = i[36].value

                bond_main_item = (c_id, name, is_listed, uscc, organ_num, legal_name, money, 
                    busin_scope, reg_address, abstract, main_bussiness)

                bond_mains.append(bond_main_item)

    wb.close()

    return bond_infos, bond_mains


def bond_info_to_mysql(conn, bond_infos, bond_mains):
    cursor = conn.cursor()

    try:
        sql = """
        insert into b_bond
        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        cursor.executemany(sql, bond_infos)

        sql = """
        insert into c_client_bond (c_id, name, is_listed, uscc, organ_num, legal_name, 
            money, busin_scope, reg_address, abstract, main_bussiness)
        values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        cursor.executemany(sql, bond_mains)

        conn.commit()
    except pymysql.DatabaseError as e:
        conn.rollback()
        raise e
    finally:
        cursor.close()

if __name__ == '__main__':
    conn = pymysql.connect(host='localhost', port=3306, user='root', passwd='root', db='financing_with_bonds', charset='utf8')
    
    # 债券信息
    files = glob.glob('g:/bond_infos/*.xlsx')
    for i in files:
        print(i)
        bond_infos, bond_mains = read_bond_info(i)
        bond_info_to_mysql(conn, bond_infos, bond_mains)

    # 债券评级信息
    files = glob.glob('g:/bond_ratings/*.xlsx')


    conn.close()

